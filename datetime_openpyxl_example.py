import datetime
from openpyxl import Workbook

wb=Workbook()

ws=wb.active

ws['A1']=42

ws.append([1,2,3])

ws['A2']=datetime.datatime.now()

wb.save("sample.xlsx")
